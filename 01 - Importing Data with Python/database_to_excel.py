# Import modules, classes, and auxiliary functions
import os
import psycopg2 as p
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from psycopg2 import Error

# Read file that contains the SQL query
with open('query.sql') as sql:
    query = sql.read()
    
try:
	# Instantiate the connection
	conn = p.connect(
	    user = os.environ['DB_USER'],
      	password = os.environ['DB_PASS'],
      	host = 'localhost',
      	port = '5432',
      	database = 'nba'
	)
	
	# Create cursor, execute the query, and fetch results
	cursor = conn.cursor()
	cursor.execute(query)
	result = cursor.fetchall()
	
	# Create workbook and select active sheet
	wb = Workbook()
	ws = wb.active
	
	# Rename active sheet
	ws.title = 'Teams'
	
	# Column headings
	column_headings = [
						'Name', 
						'City', 
						'Conference', 
						'Rank', 
						'Players', 
						'Coach', 
						'Home wins', 
						'Away wins'
	]
	ws.append(column_headings)
	
	# Add players
	for team in result:
	    ws.append(list(team[0].values()))
		
	# Get coordinates of last cell
	last_cell = ws.cell(row = ws.max_row, column = ws.max_column).coordinate
	
	# Create table
	team_table = Table(displayName = 'TeamTable', ref = 'A1:{}'.format(last_cell))
	
	# Add 'Table Style Medium 6' style
	style = TableStyleInfo(name = 'TableStyleMedium6', showRowStripes = True)
	
	# Apply style to table
	team_table.tableStyleInfo = style
	
	# Add table to spreadsheet
	ws.add_table(team_table)
	
	# Save spreadsheet
	wb.save('teams.xlsx')
	
except p.Error as error:
	print('There was an error with the database operation: {}'.format(error))
except:
    print('There was an unexpected error of type {}'.format(sys.exc_info()[0]))
finally:
	if conn:
		cursor.close()
		conn.close()